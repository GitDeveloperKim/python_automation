import openpyxl, pprint
from openpyxl.cell import *
from openpyxl.styles import Font

'''
1. openpyxl 모듈을 가져온다 
2. openpyxl.load_workbook() 함수를 호출하여 Workbook객체를 가져온다 
3. get_active)sheet() 또는 get_sheet_by_name() 메소드 호출 
4. 인덱스를 사용하거나 cell() 시트 메소드를 row와 column 키워드 매개변수와 사용
5. cell 객체를 가져와 value 속성을 읽는다 
'''

# 실습 예제 1 example.xlsx
wb = openpyxl.load_workbook('data/example.xlsx') # openpyxl 모듈을 가져와 workbook객체를 가져온다
print(wb.get_sheet_names()) # 시트 이름 얻기
sheet1 = wb.get_sheet_by_name('Sheet1') # 시트 이름으로 가져옴
print(sheet1['A1'].value)   # 데이터 얻는방법 첫번째 방법

# 데이터를 얻는 방법 두번째 방법
for i in range (1, 8, 2):
    print(i, sheet1.cell(row=i, column=2).value)

# 튜플에 저장 하기
temp = tuple(sheet1['A1':'C3'])
for rowOfCellObjects in temp:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)

# 실습 예제 2 censuspopdata.xlsx
wb = openpyxl.load_workbook('data/censuspopdata.xlsx')
sheet_census = wb.get_sheet_by_name('Population by Census Tract') # 시트 이름으로 가져옴
countyData = {}

# 통계 데이터를 가져온다
for row in range(2, sheet_census.max_row + 1):
    state = sheet_census['B'+str(row)].value
    county = sheet_census['C'+str(row)].value
    pop = sheet_census['D'+str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

# 실습 예제 3 : 액셀 쓰기
print('write results...')
resultFile = open('data/census2020.txt', 'w')
resultFile.write('allData=' + pprint.pformat(countyData))
resultFile.close()
print('Done')

# 실습 예제 4 : update example
wb = openpyxl.load_workbook('data/produceSales.xlsx')
sheet_produce = wb.get_sheet_by_name('Sheet')

PRICE_UPDATE = {'Garlic':3.07, 'Celery': 1.19, 'Lemon' : 1.27}

for rowNum in range (2, sheet_produce.max_row):
    produceName = sheet_produce.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATE:
        sheet_produce.cell(row=rowNum, column=2).value = PRICE_UPDATE[produceName]

wb.save('updateProduceSales.xlsx')

# math
